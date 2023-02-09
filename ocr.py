#-*- coding:utf-8 -*-
import re
import time
from typing import List
import requests
import os
from selenium.webdriver.common.by import By
from rpalib import excel, browser
from flows.flow import ListDataFlow, FileItem, ScriptDef, SelectItem, UserSelectItem, DirectoryItem
from rpalib import log
from rpalib.browser import Browser
from openpyxl import load_workbook
import openpyxl
logger = log.logger
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import Workbook
import datetime, time, dateutil
from dateutil.relativedelta import relativedelta
from openpyxl.styles import *

class CXlAutofit():
    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    def get_num_colnum_dict(self):
        '''
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        '''
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    # 自适应列宽
    def style_excel(self, excel_name: str, sheet_name: str):
        '''
        :param sheet_name:  excel中的sheet名
        :return:
        '''
        # 打开excel
        wb = openpyxl.load_workbook(excel_name)
        # 选择对应的sheet
        sheet = wb[sheet_name]
        # 获取最大行数与最大列数
        max_column = sheet.max_column
        max_row = sheet.max_row

        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
        num_str_dict = self.get_num_colnum_dict()

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(row=j, column=i).value
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 2.0
                    else:
                        column += 2.5
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value
        print("成功")
        # 保存
        wb.save(excel_name)

    # 调用方法 实例化类









class NB_LiXi(ListDataFlow):

    def __init__(self, summary_statement, excel_path, save_path, select_data):

        self.summary_statement = summary_statement
        self.excel_path = excel_path
        self.save_path = save_path
        self.select_data = select_data
        self.types = ""
        self.file_name = None
        self.g = None
        self.dkr_name = None
        self.dkr_id = None
        self.dkr_dkzh = None
        self.all_file = None
        self.listes = []
        self.appen = []
        self.dkr_zhlist = []


    def read_summary_statement(self):
        logger.info("开始读取表格！")
        df = pd.read_excel(self.summary_statement)
        self.g = df.groupby(['借款人姓名', '身份证号'])

    def read_ben_xi(self, excel, rowq, e, yqlv):

        # 读取第二张表
        wb = openpyxl.load_workbook(self.excel_path)
        sheets = wb.sheetnames[0]
        ws = wb[sheets]
        rows = ws.rows
        lists = []
        alls = []
        sss = [['贷款账号'], ['期数'], ['应还本息'], ['本金'], ['利息'], ['实还本金'], ['实还利息']]
        for age in sss:
            alls.append(age)
        for row in rows:
            line = [col.value for col in row]
            lists.append(line)
        yhbj_sum = 0
        bj_sum = 0
        lx_sum = 0
        shbj_sum = 0
        shlx_sum = 0
        jz_data = None
        ys_year = None
        for i in lists[1:2][0:]:
            print(i)
            print(1,str(i[1:][0]))
            print(2,str(e))
            if str(i[0:1][0]) == str(e):
                alls.append(i[0:1])
                year = str(i[1:2][0])
                ys_year = year
                year = re.search("(.*?) ", year).group(1).replace("-", "/")
                alls.append([year])
                jz_data = year
                alls.append(i[2:3])
                yhbj_sum = yhbj_sum + float(i[3:4][0])
                alls.append(i[3:4])
                bj_sum = bj_sum + float(i[4:5][0])
                alls.append(i[4:5])
                lx_sum = lx_sum + float(i[5:6][0])
                alls.append(i[5:6])
                shbj_sum = shbj_sum + float(i[5:6][0])
                alls.append(i[6:7])
                shlx_sum = shlx_sum + float(i[6:7][0])
            else:
                continue
        # print(alls)
        # 应还本金合计
        yhbj_sum = '%.2f' %yhbj_sum
        # 本金合计
        bj_sum = '%.2f' %bj_sum
        # 利息合计
        lx_sum = '%.2f' %lx_sum
        # 实还本金合计
        shbj_sum = '%.2f' %shbj_sum
        # 实还利息合计
        shlx_sum = '%.2f' %shlx_sum
        hj = [[''], ["合计"], [yhbj_sum], [bj_sum], [lx_sum], [shbj_sum], [shlx_sum]]
        for age in hj:
            alls.append(age)
        # jz_data  截止日期
        # 截止本金
        jz_bjs = float(bj_sum) - float(shbj_sum)
        jz_bj = '%.2f' %jz_bjs
        # 截止利息
        jz_lxs = float(lx_sum) - float(shlx_sum)
        jz_lx = '%.2f' %jz_lxs
        # print("截止利息", jz_lx)
        jz = [[""], [f"截止{jz_data}"], [""], [jz_bj], [jz_lx], [""], [""]]
        for abort in jz:
            alls.append(abort)
        # 最后一行
        # print(ys_year)
        datetime7 = datetime.datetime.strptime(ys_year, "%Y-%m-%d %H:%M:%S")
        zh_data = datetime7 + dateutil.relativedelta.relativedelta(days=1)  # 减1天
        zh_datas = re.search("(.*?) ", str(zh_data)).group(1).replace("-", "/")
        select_data = str(self.select_data) + " " + "00:00:00"
        srsj = datetime.datetime.strptime(f"{select_data}", "%Y-%m-%d %H:%M:%S")
        qkts = srsj - datetime7
        qktss = re.search("(.*?) ", str(qkts)).group(1)
        data = str(self.select_data).replace("-", "/")
        zh_lv = float(jz_bj) * float(yqlv) * 0.01/360*int(qktss)
        zh_lvs = '%.2f' %zh_lv
        zhs = [[""], [f"{zh_datas}-{data}"], [""], [jz_bj], [zh_lvs], [""], [""]]
        for abort_data in zhs:
            alls.append(abort_data)
        new_x = [l[0] for l in alls]
        all_list = []
        for one_list in [new_x[i:i + 7] for i in range(0, len(new_x), 7)]:
            all_list.append(one_list)
            # print(one_list)
        self.all_file = all_list

        # print(self.all_file)

        wq = excel.active
        # ws = excel.active  # 选取当前sheet工作表
        # ws.merge_cells("A1:G2")
        # # 也可以用 ws.merge_cells("A2:D2")
        # ws["A1"].alignment = Alignment(horizontal="center", vertical="center")  # 添加样式
        # ws["A1"].value = f"{self.dkr_name}利息清单"  # 添加内容
        x = ["A", "B", "C", "D", "E", "F", "G"]
        # print(rowq)
        e = int(rowq) + 2
        border = Border(left=Side(border_style=None,
                                  color='FF000000',
                                  style='thin'),
                        right=Side(border_style=None,
                                   color='FF000000',
                                   style='thin'),
                        top=Side(border_style=None,
                                 color='FF000000',
                                 style='thin'),
                        bottom=Side(border_style=None,
                                    color='FF000000',
                                    style='thin'),

                        )
        for q in self.all_file:
            # print("这是q",q)
            e += 1
            for r in range(7):

                wq[f"{x[r]}{e}"] = q[r]
                wq[f"{x[r]}{e}"].border = border
        print('这是加入小计前的行书',e)


        line = e+2
        wq[f"D{line}"].value = "小计"
        wq[f"E{line}"].value = "本金"
        wq[f"F{line}"].value = f"{jz_bj}"
        wq[f"E{int(line) + 1}"].value = "期内利息"
        wq[f"F{int(line) + 1}"].value = f"{jz_lx}"
        wq[f"E{int(line) + 2}"].value = "罚息"
        wq[f"F{int(line) + 2}"].value = f"{zh_lvs}"
        wq[f"E{int(line) + 3}"].value = "总计"
        zong_ji = float(jz_bj) + float(jz_lx) + float(zh_lvs)
        zong_ji = '%.2f' %zong_ji
        wq[f"F{int(line) + 3}"].value = f"{zong_ji}"

        # print(type(wq))
        # wr = openpyxl.load_workbook(excel)
        # sheet = wr.worksheets[0]
        a = wq.max_row
        # excel.save(os.path.join(self.save_path, f"{self.dkr_name}-{self.dkr_id}-{self.dkr_dkzh}.xlsx"))  # 保存文件
        # print(a)



        return a, bj_sum, jz_bj, jz_lx, zh_lvs

    def selet(self):
        for name, group in self.g:
            df1 = pd.DataFrame(group)
            out_excel = Workbook()

            current_row = 0
            zdkbj = 0
            # 本金
            bj = 0
            qnlx = 0
            yqfx = 0
            dkbhs = []
            for i in df1.index:

                self.dkr_name = df1["借款人姓名"][i]
                self.dkr_id = df1["身份证号"][i]
                self.dkr_dkzh = df1["贷款账号"][i]
                dkbhs.append(str(self.dkr_dkzh))
                self.dkr_zhlist.append(str(self.dkr_dkzh))

                hkfs = df1["还本付息方式"][i]
                yqlv = df1["逾期执行利率"][i]
                ws = out_excel.active  # 选取当前sheet工作表
                ws.merge_cells("A1:G2")
                # 也可以用 ws.merge_cells("A2:D2")
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")  # 添加样式
                border = Border(left=Side(border_style=None,
                                          color='FF000000',
                                          style='thin'),
                                right=Side(border_style=None,
                                           color='FF000000',
                                           style='thin'),
                                top=Side(border_style=None,
                                         color='FF000000',
                                         style='thin'),
                                bottom=Side(border_style=None,
                                            color='FF000000',
                                            style='thin'))
                ws["A1"].value = f"{self.dkr_name}利息清单"  # 添加内容
                ws["A1"].border = border
                if hkfs == "等额本息":
                    print(self.dkr_name)
                    a = self.read_ben_xi(out_excel, current_row, self.dkr_dkzh, yqlv)
                    current_row = a[0]
                    zdkbj = float(zdkbj) + float(a[1])
                    zdkbj = '%.2f' %zdkbj
                    bj = float(bj) + float(a[2])
                    bj = '%.2f' % bj
                    qnlx = float(qnlx) + float(a[3])
                    qnlx = '%.2f' % qnlx
                    yqfx = float(yqfx) + float(a[4])
                    yqfx = '%.2f' % yqfx
                    # print('money', a[1])
                    # print(zdkbj)

                    current_rows = a
                else:
                    continue

            border = Border(left=Side(border_style=None,
                                     color='FF000000',
                                     style='thin'),
                           right=Side(border_style=None,
                                      color='FF000000',
                                      style='thin'),
                           top=Side(border_style=None,
                                    color='FF000000',
                                    style='thin'),
                           bottom=Side(border_style=None,
                                       color='FF000000',
                                       style='thin'))

            # print(f"{current_rows}")
            current_row = current_rows[0]
            num = int(current_row) + 3
            # print(num)
            ws[f"A{num}"].value = ""
            ws[f"A{num}"].border = border
            ws[f"B{num}"].value = "总贷款本金"
            ws[f"B{num}"].border = border
            ws[f"C{num}"].value = "实还本金"
            ws[f"C{num}"].border = border
            ws[f"D{num}"].value = "剩余本金"
            ws[f"D{num}"].border = border
            ws[f"E{num}"].value = "剩余期内利息"
            ws[f"E{num}"].border = border
            ws[f"F{num}"].value = "剩余逾期罚息"
            ws[f"F{num}"].border = border
            ws[f"G{num}"].value = "合计欠款"
            ws[f"G{num}"].border = border

            num = int(current_row) + 4
            # print(num)
            shbj = float(zdkbj) - float(bj)
            shbj = '%.2f' % shbj
            hjqk = float(bj) + float(qnlx) + float(yqfx)
            hjqk = '%.2f' % hjqk
            ws[f"A{num}"].value = f"客户总计（截止{self.select_data}）"
            ws[f"A{num}"].border = border
            ws[f"B{num}"].value = f"{zdkbj}"
            ws[f"B{num}"].border = border
            ws[f"C{num}"].value = f"{shbj}"
            ws[f"C{num}"].border = border
            ws[f"D{num}"].value = f"{bj}"
            ws[f"D{num}"].border = border
            ws[f"E{num}"].value = f"{qnlx}"
            ws[f"E{num}"].border = border
            ws[f"F{num}"].value = f"{yqfx}"
            ws[f"F{num}"].border = border
            ws[f"G{num}"].value = f"{hjqk}"
            ws[f"G{num}"].border = border
            out_excel.save(os.path.join(self.save_path, f"{self.dkr_name}-{self.dkr_id}.xlsx"))
            a = os.path.join(self.save_path, f"{self.dkr_name}-{self.dkr_id}.xlsx")
            # print(dkbhs)
            self.formatting(dkbhs, a)


    def formatting(self, num, path):
        for i in num:

            file = os.path.join(self.save_path, path)


            ex1 = load_workbook(file)
            sheet = ex1.active
            sheets = ex1.sheetnames
            ws = ex1[sheets[0]]
            zhs = ws.max_row
            # print("总行数", zhs)
            lists = []
            for row in sheet.iter_rows(min_row=1, max_row=51104,
                                       min_col=1, max_col=2):
                for cell in row:
                    if cell.value == i:
                        # print(cell.coordinate)
                        lists.append(cell.coordinate)
            try:
                # print(lists)
                a = lists[0]
                b = lists[-1]
                sheet.merge_cells(f"{a}:{b}")
                alignment_center = Alignment(horizontal='center', vertical='center')
                # 指定区域单元格居中
                # print(f"{zhs}")
                ws_area = ws[f"A3:G{zhs}"]
                for i in ws_area:
                    for j in i:
                        j.alignment = alignment_center
                ex1.save(file)
                #
                # list_xj = []
                # for row in sheet.iter_rows(min_row=1, max_row=51104,
                #                            min_col=1, max_col=8):
                #     for cell in row:
                #         if cell.value == "小计":
                #             # print("小计", cell.coordinate)
                #             list_xj.append(cell.coordinate)
                #
                # list_xj1 = []
                # x = None
                # for i in list_xj:
                #     c = re.search("D(.*)", i).group(1)
                #     list_xj1.append(c)
                # # print(111,list_xj1)
                # #
                #
                # list_xjs = []
                #
                #
                #
                # for row in sheet.iter_rows(min_row=1, max_row=51104,
                #                            min_col=1, max_col=2):
                #     for cell in row:
                #         if cell.value == "贷款账号":
                #             # print("小计", cell.coordinate)
                #             list_xjs.append(cell.coordinate)
                #
                # list_xj2 = []
                # x = None
                # for i in list_xjs:
                #     c = re.search("A(.*)", i).group(1)
                #     list_xj2.append(c)
                # # print(222,list_xj2)
                # print(len(list_xj2))
                # for i in len(list_xj2):
                #
                #     self.format_border("sheet", list_xj2[i], list_xj1[i], 1, 7)












                # c = re.search("A(.*)", lists[0]).group(1)
                #
                # try:
                #     hangshu = int(c) + 4
                #
                # except:
                #     hangshu = int(c) - 3
                # self.format_border("sheet", 1, hangshu, 1, 7)




            except:
                pass
        # self.format_border("sheet", 1, e, 1, 7)





    def run(self):
        self.read_summary_statement()
        self.selet()
        for i in os.listdir(self.save_path):
            if ".xlsx" in i:
                path = os.path.join(self.save_path, i)
                Entity = CXlAutofit()
                Entity.style_excel(path, 'Sheet')


        # print(self.appen)
        # print(self.dkr_zhlist)
        # print(self.listes)
        # self.formatting()



export = ScriptDef(
    cls=NB_LiXi,
    group="其他",
    title="宁波银行证据合并",
    arguments=[FileItem(title="excel表格", name="excel_file"),
               SelectItem(title="选择功能", name="SFunction", options=["民特卷宗归档", "民初卷宗归档"]),

    ]
)


if __name__ == '__main__':
    a = NB_LiXi(r"C:\Users\9000\Desktop\测试\汉资100户清单20220607.xlsx", r"C:\Users\9000\Desktop\测试\等额本息-20220607.xlsx",
                r"C:\Users\9000\Desktop\测试", r"2020-12-3").run()
