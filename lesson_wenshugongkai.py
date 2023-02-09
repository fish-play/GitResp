import os
import re
import time
import random
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook, Workbook
from rpalib.log import logger
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from flows.flow import ScriptDef, FileItem, SelectItem, UserSelectItem, ListDataFlow,DirectoryItem
from rpalib.log import logger
from rpalib import excel, browser


class Item(object):
    xxx: str  # 案号


class Wenshugongkai(ListDataFlow):

    def __init__(self, "要用到的参数"):
        super().__init__()
        self."用到的参数" = "用到的参数"
        ...


    def login(self):
        self.b = browser.create_IE()  # 启动浏览器驱动
        self.b.load(f"xxxxx")  # 打开网站
        self.b.clear_input('选中输入框', self.username)  # 输入用户名
        self.b.clear_input('选中输入框', self.password)  # 输入密码
        self.b.click('登录按钮')  # 点击登录
        self.b.is_not_visible('登录按钮')  # 一直等待某个元素消失，默认超时10秒
        self.b.load(url='')  # 输入网址
        self.b.load(
            url=''
        )
        self.b.switch_to_frame(xpath='')  # 定位iframe标签元素
        self.b.click(xpath='')  # 点击网上办理

    def chaxun_sousuo(self, item):
        self.b.switch_to_window_by_title(title='')  # 切换表头
        self.b.switch_to_frame()  # 定位iframe
        self.b.switch_to_frame()  # 定位iframe
        self.b.clear_input(xpath='', data=f'{item.an_hao}\n')  # 输入...
        time.sleep(2)
        try:
            self.b.click(xpath=f'//a[contains(@onclick,"{item.an_hao}")][text()="上网办理"]')  # 点击...
        except:
            logger.info(f'案号：{item.an_hao}，没有查询到，跳过该案号')  # 打印日志
            return False






    def daoruwenshu(self, "用到的参数(item)"):
        self.b.switch_to_window_by_title(title='')  # 切换页面头信息
        self.b.switch_to_frame(xpath='')  # 定位iframe
        try:
            flag1 = self.b.find_element(xpath='').get_attribute('title')  # 获取图内所有的title信息
            flag2 = self.b.find_element(xpath='').get_attribute('title')  # 获取图内所有的title信息
            if flag1 == "处理成功" and flag2 == "处理成功":
                self.b.default_content()  # 移除页面焦点
                self.b.click('')  # 点击其他要求
                return False
            self.b.click(xpath=f'//div[@id="divMark"]//div[text()="{item.an_hao}"]/../..//a[text()="文书办理"]',
                         wait_time=3')  # 点击什么东西
            logger.info(f'案号：{item.an_hao}--导入文书已做过')  # 打印日志
        except:
            logger.info(f'{item.an_hao}--文书没有导入--开始导入文书')  # 打印日志
            self.b.click(xpath="")  # 点击导入文书
            self.b.default_content()  # 移除页面焦点
            self.b.switch_to_frame(xpath='')  # 选中iframe元素
            file, file_path = self.get_file_path'''下边写的函数'''(self.words_dir'''class中给的参数''', item.an_hao)  # 文件名和文件路径
            logger.info(f'文件名：{file}')  # 日志打印文件名
            logger.info(f'文件路径：{file_path}')  # 日志打印文件路径
            logger.info(f'------------------------------------------------------')  # 日志打印分隔线
            self.b.fill_input(xpath='' '''选中要上传的文书''',value=file_path)  # 选中要上传的文书，并提供路径
            self.b.driver.find_elements(By.XPATH, '')  # 找到所有的...
            assert file in self.b.driver.page_source   # 断言，file是否在页面源码中，如果在则执行下一步，不在报错
            self.b.default_content()  # 移除页面焦点
            self.b.click(xpath='')  # 点击下一步操作
            logger.info('---文书上传完成---')
            self.b.default_content()  # 移除页面焦点
            self.b.switch_to_frame(xpath='')  # 选中iframe

    def baocun_tijiao(self, "用到的参数（item）"):
        for _ in range(10):  # 遍历10次
            logger.info('---进入到保持页面---')
            self.b.switch_to_window_by_title(title='')  # 切换页面表头
            self.b.switch_to_frame(xpath='', sleep_time=3)  # 定位iframe，等待3秒
            self.b.click(xpath=f'//div[@id="divMark"]//div[text()="{item.an_hao}"]/../..//a[text()="文书办理"]', wait_time=3)  # 点击文书办理，静态等待3秒
            self.b.driver.switch_to.default_content()  # 移除焦点
            self.b.switch_to_frame(xpath='')  # 定位iframe
            try:
                self.b.find_element(xpath='')
                print('找不到')
                self.b.default_content()  # 移除焦点
                self.b.click(xpath='')
                logger.info('----------保持失败----------')
                self.wenshuchushihua()
                continue
            except:
                print("找到")
                self.b.default_content()  # 移除焦点
                self.b.switch_to_frame(xpath='')  # 定位iframe
                button = self.b.click(xpath='', wait_time=50)  # 定位元素
                self.b.driver.execute_script('', button)  # 获取 JavaScript 执行后的返回值的
                break

    def wenshuchushihua(self):
        self.b.switch_to_window_by_title(title='')  # 切换页面表头
        self.b.switch_to_frame(xpath='')  # 定位iframe
        self.b.find_element(xpath='')  # 获取元素
        ActionChains(self.b.driver).click(
            self.b.driver.find_element(By.XPATH, '')
        ).perform()  # 链式鼠标左键点击该元素
        self.b.click(xpath='', sleep_time=2)  # 点击文书初始化
        self.b.default_content()  # 移除焦点
        self.b.click(xpath='')
        self.b.click(xpath='')  # 点击确定
        logger.info('---初始化完成---')

    def chongxinfenxi_baocuntijiao(self):
        self.b.switch_to_window_by_title(title='')  # 切换表头
        self.b.switch_to_frame(xpath='')  # 定位iframe
        button = self.b.find_element(xpath='')  # 点击重新分析
        self.b.driver.execute_script("", button)  # 获取 JavaScript 执行后的返回值的
        try:
            self.b.switch_to_alert().accept()  # 点击弹窗确认按钮
        except Exception as e:
            print('报错')
            print(e)
        self.b.switch_to_window_by_title(title='')  # 切换表头
        self.b.switch_to_frame(xpath='')  # 定位iframe
        self.b.click(xpath='')  # 点击什么
        self.b.switch_to_frame()  # 定位iframe
        self.b.click(xpath='')  # 点击...
        self.b.click(xpath='')  # 点击...
        self.b.driver.switch_to.parent_frame()  # 切换到第一个iframe
        self.b.default_content()  # 取消iframe定位
        self.b.switch_to_frame(xpath='')  # 定位iframe
        button = self.b.find_element('')  # 点击提交
        self.b.driver.execute_script('', button)  # 获取 js 执行后返回的值
        self.b.default_content()  # 取消iframe定位
        self.b.click(xpath='', wait_time=2)  # 点击..., 隐式等待2秒
        self.b.default_content()  # 取消iframe定位
        self.b.click(xpath='')  # ...

        # 加载数据
    def read_data(self):
        data = excel.loads(self.excel_file'''class里给的excel参数''', Item, {
            "anz-hao": {"title": "案号", "required": True},
        })       # 加载excel表中的数据
        return data

    # 写入数据
    def write_data(self, process_list):
        wb_save = Workbook('执行情况')  # 创建workbook对象 生成表单
        sh1 = wb_save.create_sheet()
        sh1.append(["案号", "执行情况"])  # 添加案号和执行情况
        for process in process_list:  # 遍历process_list
            sh1.append(process)  # 循环添加到process
        f = os.path.join(os.getcwd(), "程序执行情况")  # 格式当前文件路径
        wb_save.save(f)  # 保存
        logger.info(f"执行情况保存在{f}")

    # 运行程序
    def run(self):
        # 加载表格，读取数据
        process_list = []
        data = self.read_data()  # 调用read_data 加载数据
        self.login()  # 调用login 登录
        try:
            for item in data:  # 循环加载数据函数
                try:
                    flag = self.chaxun_sousuo(item)  # 掉查询搜索的item
                    if flag is False:
                        continue
                    flag = self.daoruwenshu(item)  # 成功后调用导入文书
                    if flag is False:
                        continue
                    self.baocun_tijiao(item)  # 成功保存
                    self.chongxinfenxi_baocuntijiao()
                except:
                    process_list.append([item.an_hao, '失败'])
                    self.b.quit()
                    self.login()
                    continue
        finally:
            if len(process_list):   # 如果长度不为空
                from openpyxl import load_workbook, Workbook
                wb_save = Workbook('1')
                sh1 = wb_save.create_sheet()
                sh1.append([
                    "案号", "执行情况"
                ])
                for e in process_list:
                    sh1.append(e)
                f = os.path.join(os.getcwd(), "执行情况.xlsx")  # 格式当前文件路径
                wb_save.save(f)
                logger.info(f'执行情况保存在{f}')



export = ScriptDef(
    cls = Wenshugongkai,
    group = '省份-地区-文书公开',
    title = '省份-地区-文书公开',
    description = '省份-地区-文书公开',
    arguments=[

        UserSelectItem(title='账号', username_field='name', password_field='pwd'),
        FileItem(title='excel表格', name='excel_file'),
        DirectoryItem(title='裁定书路径', name='words_dir'),
        SelectItem(title='目标网址', name='url', options=[
            r'xxxxxxxxxxx'  # 网址
        ])

    ]

)


# 沙区运行
if __name__ == '__main__':
    Wenshugongkai(
        url=r'http://www.baidu.com',   # 开头给的参数地址
        name='zhaofengxiang',    # 开头给的参数姓名
        pwd='2132131',      # # 开头给的参数密码
        excel_file=r'',    # # 开头给的参数excel表的地址
        words_dir=r''     # 开头给的参数...
    ).run()
