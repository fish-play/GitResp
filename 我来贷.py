# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
import shutil

from rpalib.log import logger
from flows.flow import ScriptDef, DirectoryItem, FileItem



def get_changeinfo(filename):
    """
    根据表头,获取信息并核对
    :param filename: 表头.excel
    :return:
    """
    wb = load_workbook(filename)
    ws = wb.active

    incom_num, names, idcards, national, household_address, contracts = ws['A'], ws['B'], ws['G'], ws['E'], ws['J'], ws[
        'O']

    sign_times, start_times, end_times, npers, rates, contract_amounts = ws['P'], ws['Q'], ws['R'], ws['T'], ws['U'], \
                                                                         ws['V']
    # 身份证
    id_card_list = [(idcards[idcard].value) for idcard in range(len(idcards))]
    # 名称
    name_list = [(names[name].value) for name in range(len(names))]
    # 进件号
    incom_list = [(incom_num[incom].value) for incom in range(len(incom_num))]
    # 民族
    national_list = [(national[nation].value) for nation in range(len(national))]
    # 户籍地址
    household_list = [(household_address[household].value) for household in range(len(household_address))]
    # 合同编号
    contract_list = [(contracts[contract].value) for contract in range(len(contracts))]
    # 签署合同日期
    singn_time_list = [(sign_times[sign_time]).value for sign_time in range(len(sign_times))]
    singn_time_list = [(str(sign_time).split(" ")[0]) for sign_time in singn_time_list]
    # 借款起始日
    start_time_list = [(start_times[start_time].value) for start_time in range(len(start_times))]
    start_time_list = [(str(start_time).split(" ")[0]) for start_time in start_time_list]
    # 借款截止日
    end_time_list = [(end_times[end_time].value) for end_time in range(len(end_times))]
    end_time_list = [(str(end_time).split(" ")[0]) for end_time in end_time_list]
    # 贷款期数
    nper_list = [(npers[nper].value) for nper in range(len(npers))]
    # 年利率
    rate_list = [(rates[rate].value) for rate in range(len(rates))]
    # 合同金额
    contract_amount_list = [(contract_amounts[contract_amount].value) for contract_amount in
                            range(len(contract_amounts))]
    # infos = set(zip(name_list, id_card_list, incom_list))

    infos = set(zip(
        # 姓名列表,身份证列表,进件号列表
        name_list, id_card_list, incom_list,
        # 民族列表,户籍地址列表
        national_list, household_list,

        # 合同编号, 贷款金额, 期数, 借款起始日, 借款截止日, 借款年利率, 合同签署日期
        # 合同编号,借款金额列表,贷款期数
        contract_list, contract_amount_list, nper_list,
        # 借款起始日列表,借款截止日列表
        start_time_list, end_time_list,
        # 年利率  合同签署日期
        rate_list, singn_time_list
    ))

    return infos


def Loan_voucher(floders, output, infos):
    """

    :param floders: 放款凭证文件夹
    :param output: 重命名文件夹目录
    :param infos:
    :return:
    """
    logger.info(f"开始移动放款凭证")

    try:
        for info in infos:
            # print(i[1])
            if info[0] == '姓名':
                continue
            xm = info[0]
            if info[1] == "身份证号":
                continue
            sfz = info[1]
            if info[2] == '进件编号':
                continue
            name = os.path.join(floders, info[2])
            a = os.listdir(name)
            filename = f'放款凭证-{xm}-{sfz}.pdf'
            for e in a:
                if e == filename:
                    # print(e)
                    wj_name = e
                    rfloder = f'{info[0]}{info[1]}'
                    if not os.path.exists(os.path.join(output, rfloder)):
                        os.makedirs(os.path.join(output, rfloder))
                    print(wj_name)
                    print(os.path.join(name, wj_name))
                    shutil.copyfile(os.path.join(name, wj_name), os.path.join(output, rfloder, wj_name))
                    print("移动完成")
                    logger.info(f'{filename} 移动完成')
        logger.info(f'文件保存在{output}')
    except Exception as e:
        logger.error(e)


def run(dataheader, processfloder, outputfloder):
    infos = get_changeinfo(dataheader)
    Loan_voucher(processfloder, outputfloder, infos)

#
# run(r'C:\Users\9000\Desktop\我来贷-数据表头（三个程序通用表头）.xlsx',
#     r'C:\Users\9000\Desktop\new样例',
#     r'C:\Users\9000\Desktop\tmp')

export = ScriptDef(
    func=run(),
    group="其他",
    title="我来贷放款凭证处理",
    arguments=[
        FileItem(title="我来贷数据表头", name="dataheader"),
        DirectoryItem(title="处理文件夹目录",name="processfloder"),
        DirectoryItem(title="输出文件夹目录",name="outputfloder")
    ]
)
